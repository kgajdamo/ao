"""
Weekly test progress chart for MoE Training tests on XPU.
Reads data from test_progress_data.xlsx and generates a line chart
showing passing test rate (%) per test file over time.

Excel layout:
  Sheet "totals":             test_file | total
  Sheet "weekly_pass_counts": date | file1 | file2 | ...
"""

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(__file__).parent / "test_progress_data.xlsx"

# ── Load data from Excel ─────────────────────────────────────────────

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# Read totals
ws_totals = wb["totals"]
rows = list(ws_totals.iter_rows(min_row=2, values_only=True))  # skip header
test_totals = {row[0]: row[1] for row in rows}

# Read weekly pass counts
ws_weekly = wb["weekly_pass_counts"]
header = [cell.value for cell in next(ws_weekly.iter_rows(min_row=1, max_row=1))]
test_files = header[1:]  # first column is "date"

dates = []
weekly_pass_counts = []  # list of dicts
for row in ws_weekly.iter_rows(min_row=2, values_only=True):
    dates.append(datetime.strptime(str(row[0]), "%Y-%m-%d"))
    weekly_pass_counts.append(
        {test_files[i]: row[i + 1] for i in range(len(test_files))}
    )

wb.close()

# ── Compute pass rates and plot ──────────────────────────────────────


def pass_rate(pass_count, test_file):
    total = test_totals[test_file]
    return (pass_count / total * 100) if total > 0 else 0.0


fig, ax = plt.subplots(figsize=(14, 8))
fig.suptitle(
    "MoE Training XPU Tests — Weekly Pass Rate (%)",
    fontsize=14,
    fontweight="bold",
)

colors = plt.cm.tab20.colors

for idx, test in enumerate(test_files):
    values = [pass_rate(w[test], test) for w in weekly_pass_counts]
    ax.plot(
        dates,
        values,
        marker="o",
        label=test,
        color=colors[idx % len(colors)],
        linewidth=2,
    )

    # Annotate the latest point with +/- diff if it changed from the previous week
    if len(values) >= 2:
        diff = values[-1] - values[-2]
        if abs(diff) > 0.01:
            sign = "+" if diff > 0 else ""
            ax.annotate(
                f"{sign}{diff:.1f}%",
                xy=(dates[-1], values[-1]),
                xytext=(8, 0),
                textcoords="offset points",
                fontsize=7,
                fontweight="bold",
                color="green" if diff > 0 else "red",
                va="center",
            )

ax.set_ylabel("Pass Rate (%)")
ax.set_ylim(-5, 105)
ax.legend(loc="center left", bbox_to_anchor=(1.0, 0.5), fontsize=8)
ax.grid(True, alpha=0.3)
ax.xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=mdates.MO))
ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")
ax.set_xlabel("Date")

plt.tight_layout()
out_path = Path(__file__).parent / "test_progress_chart.png"
plt.savefig(out_path, dpi=150, bbox_inches="tight")
print(f"Chart saved to {out_path}")
plt.show()
